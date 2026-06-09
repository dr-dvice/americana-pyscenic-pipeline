#!/usr/bin/env python3
"""
Create formatted Excel spreadsheet summarizing all regulons:
- TF name, avg AUC, avg NES
- Target genes with expression weights and product names from GFF
"""
import argparse
import pandas as pd
import numpy as np
import re
import ast
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def main():
    parser = argparse.ArgumentParser(description='Create regulon summary Excel spreadsheet')
    parser.add_argument('--regulons', required=True, help='Regulons CSV from pySCENIC ctx')
    parser.add_argument('--gff_key', required=True, help='GFF key TSV (transcript_id, product, rna_type)')
    parser.add_argument('--orthologs', required=True, help='Species-to-fly ortholog TSV')
    parser.add_argument('--output', required=True, help='Output Excel file')
    args = parser.parse_args()

    # Load GFF key — handle both column name conventions
    print("Loading GFF key...")
    gff_key = pd.read_csv(args.gff_key, sep='\t')
    # Detect gene ID column: 'transcript_id' (old 2025) or 'Gene' (bellini2026)
    if 'transcript_id' in gff_key.columns:
        id_col = 'transcript_id'
    elif 'Gene' in gff_key.columns:
        id_col = 'Gene'
    else:
        raise ValueError(f"GFF key must have 'transcript_id' or 'Gene' column. Found: {list(gff_key.columns)}")
    gene_info = {}
    for _, row in gff_key.iterrows():
        gene_info[row[id_col]] = {
            'product': row['product'] if pd.notna(row['product']) else 'uncharacterized',
            'rna_type': row['rna_type'] if pd.notna(row['rna_type']) else 'N/A'
        }
    print(f"  Loaded {len(gene_info)} gene annotations (ID column: '{id_col}')")

    # Load ortholog mapping (best hit per gene)
    print("Loading ortholog mapping...")
    orthologs = pd.read_csv(args.orthologs, sep='\t')
    orthologs = orthologs.sort_values('bitscore', ascending=False).drop_duplicates('sp_transcript')
    ortholog_map = dict(zip(orthologs['sp_transcript'], orthologs['fly_gene']))

    # Parse regulons CSV
    print("Parsing regulons...")
    regulons = pd.read_csv(args.regulons, skiprows=1, header=[0, 1])
    regulons.columns = [col[1] if 'Unnamed' not in col[1] else col[0] for col in regulons.columns]

    # Aggregate per TF
    print("Aggregating per TF...")
    tf_data = {}
    for _, row in regulons.iterrows():
        tf_id = row['TF']
        auc = float(row['AUC'])
        nes = float(row['NES'])
        target_str = row['TargetGenes']
        try:
            targets = ast.literal_eval(target_str)
        except:
            continue
        if tf_id not in tf_data:
            tf_data[tf_id] = {'auc_scores': [], 'nes_scores': [], 'targets': {}}
        tf_data[tf_id]['auc_scores'].append(auc)
        tf_data[tf_id]['nes_scores'].append(nes)
        for gene_id, weight in targets:
            if gene_id not in tf_data[tf_id]['targets'] or weight > tf_data[tf_id]['targets'][gene_id]:
                tf_data[tf_id]['targets'][gene_id] = weight

    # Build summary rows
    print("Building summary...")
    tf_summaries = []
    for tf_id, data in tf_data.items():
        info = gene_info.get(tf_id, {'product': 'uncharacterized', 'rna_type': 'N/A'})
        fly = ortholog_map.get(tf_id, 'N/A')
        tf_summaries.append({
            'tf_id': tf_id,
            'fly_ortholog': fly,
            'tf_product': info['product'],
            'tf_rna_type': info['rna_type'],
            'avg_auc': np.mean(data['auc_scores']),
            'avg_nes': np.mean(data['nes_scores']),
            'n_motifs': len(data['auc_scores']),
            'n_targets': len(data['targets']),
            'targets': data['targets']
        })

    tf_summaries.sort(key=lambda x: x['avg_nes'], reverse=True)
    print(f"  {len(tf_summaries)} TFs, {sum(s['n_targets'] for s in tf_summaries)} total target entries")

    # Create Excel workbook
    print("Creating Excel spreadsheet...")
    wb = Workbook()

    # Styles
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    tf_header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    target_header_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    alt_row_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    thin_border = Border(bottom=Side(style='thin', color='C0C0C0'))

    # Sheet 1: Summary table
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary_headers = ['TF ID', 'Product', 'RNA Type', 'Fly Ortholog', 'Avg AUC', 'Avg NES', 'Motifs', 'Targets']
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = tf_header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(bottom=Side(style='medium', color='2F5496'))

    for i, tf in enumerate(tf_summaries):
        row = i + 2
        values = [tf['tf_id'], tf['tf_product'], tf['tf_rna_type'], tf['fly_ortholog'],
                  round(tf['avg_auc'], 4), round(tf['avg_nes'], 2), tf['n_motifs'], tf['n_targets']]
        for col, val in enumerate(values, 1):
            cell = ws_summary.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=10)
            if col in [5, 6]:
                cell.number_format = '0.0000'
                cell.alignment = Alignment(horizontal='center')
            if row % 2 == 0:
                cell.fill = alt_row_fill

    summary_widths = [18, 40, 10, 14, 10, 10, 8, 8]
    for i, w in enumerate(summary_widths, 1):
        ws_summary.column_dimensions[get_column_letter(i)].width = w

    # Per-TF sheets
    used_sheet_names = {}
    for tf in tf_summaries:
        fly = tf['fly_ortholog']
        if fly != 'N/A' and '-like' not in tf['tf_product']:
            base_name = fly
        else:
            base_name = tf['tf_id']
        base_name = re.sub(r'[/\\?*\[\]:]', '_', base_name)[:28]

        if base_name in used_sheet_names:
            used_sheet_names[base_name] += 1
            sheet_name = f"{base_name}_{used_sheet_names[base_name]}"
        else:
            used_sheet_names[base_name] = 1
            sheet_name = base_name

        ws = wb.create_sheet(title=sheet_name)

        # TF header block
        ws.cell(row=1, column=1, value=f"TF: {tf['tf_id']}")
        ws.cell(row=1, column=1).font = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
        ws.cell(row=1, column=1).fill = tf_header_fill
        ws.cell(row=1, column=1).alignment = Alignment(vertical='center')
        for col in range(1, 6):
            ws.cell(row=1, column=col).fill = tf_header_fill
            ws.cell(row=1, column=col).border = Border(bottom=Side(style='thin', color='FFFFFF'))
        ws.row_dimensions[1].height = 22

        ws.cell(row=2, column=1, value="Product:")
        ws.cell(row=2, column=1).font = Font(name='Calibri', size=10, bold=True)
        ws.cell(row=2, column=2, value=tf['tf_product'])
        ws.cell(row=2, column=2).font = Font(name='Calibri', size=10)
        ws.cell(row=2, column=3, value="RNA Type:")
        ws.cell(row=2, column=3).font = Font(name='Calibri', size=10, bold=True)
        ws.cell(row=2, column=4, value=tf['tf_rna_type'])
        ws.cell(row=2, column=4).font = Font(name='Calibri', size=10)

        ws.cell(row=3, column=1, value="Fly Ortholog:")
        ws.cell(row=3, column=1).font = Font(name='Calibri', size=10, bold=True)
        ws.cell(row=3, column=2, value=tf['fly_ortholog'])
        ws.cell(row=3, column=2).font = Font(name='Calibri', size=10)

        ws.cell(row=4, column=1, value="Avg AUC:")
        ws.cell(row=4, column=1).font = Font(name='Calibri', size=10, bold=True)
        ws.cell(row=4, column=2, value=round(tf['avg_auc'], 4))
        ws.cell(row=4, column=2).font = Font(name='Calibri', size=10)
        ws.cell(row=4, column=3, value="Avg NES:")
        ws.cell(row=4, column=3).font = Font(name='Calibri', size=10, bold=True)
        ws.cell(row=4, column=4, value=round(tf['avg_nes'], 2))
        ws.cell(row=4, column=4).font = Font(name='Calibri', size=10)
        ws.cell(row=4, column=5, value=f"Motifs: {tf['n_motifs']}  |  Targets: {tf['n_targets']}")
        ws.cell(row=4, column=5).font = Font(name='Calibri', size=10, italic=True, color='666666')

        ws.row_dimensions[5].height = 6

        target_headers = ['Gene ID', 'Product', 'RNA Type', 'Fly Ortholog', 'Expression Weight']
        for col, header in enumerate(target_headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = Font(name='Calibri', size=10, bold=True)
            cell.fill = target_header_fill
            cell.border = Border(
                top=Side(style='thin', color='2F5496'),
                bottom=Side(style='thin', color='2F5496')
            )
            cell.alignment = Alignment(horizontal='center')

        sorted_targets = sorted(tf['targets'].items(), key=lambda x: x[1], reverse=True)
        for i, (gene_id, weight) in enumerate(sorted_targets):
            row = i + 7
            info = gene_info.get(gene_id, {'product': 'uncharacterized', 'rna_type': 'N/A'})
            fly = ortholog_map.get(gene_id, 'N/A')
            values = [gene_id, info['product'], info['rna_type'], fly, round(weight, 4)]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = Font(name='Calibri', size=10)
                cell.border = thin_border
                if col == 5:
                    cell.alignment = Alignment(horizontal='center')
                    cell.number_format = '0.0000'
                if row % 2 == 0:
                    cell.fill = alt_row_fill

        target_widths = [18, 45, 10, 14, 16]
        for i, w in enumerate(target_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(args.output)
    print(f"\nSaved: {args.output}")
    print(f"  - Summary sheet: {len(tf_summaries)} TFs")
    print(f"  - Individual sheets: one per TF with sorted target gene lists")

if __name__ == '__main__':
    main()
